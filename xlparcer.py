import openpyxl
from aiogram import Bot, Dispatcher, executor, types
import logging
from openpyxl import load_workbook
import os

# constants
WEEKDAYS = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
WEEKDAYS_RUS = ['Понедельник', 'Вторник', 'Среда', 'Четверг', 'Пятница', 'Суббота', 'Воскресенье']
PERIOD = ['8:00 - 9:35', '9:50 - 11:25', '11:40 - 13:15', '14:00 - 15:35', '15:50 - 17:25', '17:40 - 19.15']
DEFAULT_TIME_COLUMN = 2

os.getcwd()     # file needs to be in the same folder as the script

# FILE_NAME only gets .xlsx files, you need to convert xls into xlsx
FILE_NAME = 'rasp.xlsx'
WORKBOOK = load_workbook(FILE_NAME)
WORKSHEET_NAMES = WORKBOOK.sheetnames


def get_time(worksheet, row):
    cell = worksheet[coord(row, DEFAULT_TIME_COLUMN)]
    time = cell.value
    return str(time)


def coord(row, column):
    s = chr(column + 64) + str(row)     # converting R1C1 into A1 notation, column number cant be higher than 26
    return s


def time2period(time):
    s = ''
    for c in time:
        s = s + c
        if s == '8':
            return 1
        elif s == '9':
            return 2
        elif s == '11':
            return 3
        elif s == '14':
            return 4
        elif s == '15':
            return 5
        elif s == '17':
            return 6


class Lesson(object):
    def __init__(self, sheet_num, row, column):
        self.time = None                # time of this period (string from xlsx table)
        self.period = None              # integer number of period
        self.sheet_num = sheet_num      # number of sheet of this lesson in xlsx file
        self.row = row
        self.column = column
        if self.row > 0 and self.column > 0 and self.sheet_num > 0:
            self.sheet = WORKBOOK[WORKSHEET_NAMES[self.sheet_num]]
            cell = self.sheet[coord(self.row, self.column)]
            self.discipline = cell.value
        else:
            print('incorrect row/column/sheet number')


class Day(Lesson):
    def __init__(self, sheet_num, row, column):     # coordinates of first period of the day
        super().__init__(sheet_num, row, column)

    def print_day(self):
        ans = []
        for period_row in range(self.row, self.row + 5):
            cell = self.sheet[coord(period_row, self.column)]
            ans.append(get_time(self.sheet, period_row) + ' ' + str(cell.value))
        return ans








